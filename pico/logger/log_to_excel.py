
import datetime
import sys
import time
import serial

from openpyxl import Workbook, load_workbook


SERIAL_PORT = "COM3"         
BAUD_RATE = 115200
EXCEL_PATH = r"C:\Users\Khoa\Desktop\sensor_log.xlsx"  


HEADER = ["timestamp", "air_quality", "lux", "temperature_c", "humidity_pct", "sound_level"]


def get_or_create_workbook(path):
    try:
        wb = load_workbook(path)
        ws = wb.active
        return wb, ws
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sensor Data"
        ws.append(HEADER)
        wb.save(path)
        return wb, ws


def main():
    print(f"Opening serial port {SERIAL_PORT} @ {BAUD_RATE}...")
    try:
        ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=1)
    except serial.SerialException as e:
        print(f"Could not open {SERIAL_PORT}: {e}")
        print("Check Device Manager for the correct COM port and update SERIAL_PORT above.")
        sys.exit(1)

    print(f"Logging to: {EXCEL_PATH}")
    wb, ws = get_or_create_workbook(EXCEL_PATH)

    print("Listening for data. Press Ctrl+C to stop.\n")
    try:
        while True:
            raw = ser.readline().decode("utf-8", errors="ignore").strip()
            if not raw:
                continue

            if raw.startswith("DATA,"):
                parts = raw.split(",")[1:]  # drop the "DATA" label
                if len(parts) != 5:
                    print(f"Skipping malformed line: {raw}")
                    continue

                timestamp = datetime.datetime.now().isoformat(timespec="seconds")
                row = [timestamp] + parts
                ws.append(row)
                wb.save(EXCEL_PATH)  # save every row so nothing is lost if it crashes

                print(f"Logged: {row}")
            else:
                # Non-data line (status messages etc.) - just echo it
                print(f"[Pico] {raw}")

    except KeyboardInterrupt:
        print("\nStopping. Final save...")
        wb.save(EXCEL_PATH)
        print("Done.")
    finally:
        ser.close()


if __name__ == "__main__":
    main()
